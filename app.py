from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os


# =====================================================
# TEACHER LOGIN SETTINGS
# =====================================================

TEACHERS = [

    {
        "email": "teacher@gmail.com",
        "password": "123456",
        "name": "Admin Teacher",
        "role": "teacher"
    },

    {
        "email": "cse@college.com",
        "password": "cse123",
        "name": "CSE Teacher",
        "role": "teacher"
    },

    {
        "email": "it@college.com",
        "password": "it123",
        "name": "IT Teacher",
        "role": "teacher"
    },

    {
        "email": "aids@college.com",
        "password": "aids123",
        "name": "AI & DS Teacher",
        "role": "teacher"
    },

    {
        "email": "entc@college.com",
        "password": "entc123",
        "name": "ENTC Teacher",
        "role": "teacher"
    }

]


# =====================================================
# FLASK APP
# =====================================================

app = Flask(__name__)

CORS(app)


# =====================================================
# EXCEL SETTINGS
# =====================================================

BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)

EXCEL_FOLDER = os.path.join(
    BASE_DIR,
    "exports"
)

if not os.path.exists(EXCEL_FOLDER):

    os.makedirs(EXCEL_FOLDER)


# =====================================================
# BRANCH EXCEL FILE
# =====================================================

def get_excel_file(class_name):

    safe_name = str(class_name).replace(
        " ",
        "_"
    )

    safe_name = safe_name.replace(
        "&",
        "and"
    )

    return os.path.join(
        EXCEL_FOLDER,
        f"{safe_name}_Attendance.xlsx"
    )


# =====================================================
# EXCEL STYLES
# =====================================================

header_fill = PatternFill(
    fill_type="solid",
    fgColor="2563EB"
)

header_font = Font(
    color="FFFFFF",
    bold=True
)

thin_border = Border(
    left=Side(
        style="thin",
        color="D1D5DB"
    ),
    right=Side(
        style="thin",
        color="D1D5DB"
    ),
    top=Side(
        style="thin",
        color="D1D5DB"
    ),
    bottom=Side(
        style="thin",
        color="D1D5DB"
    )
)


# =====================================================
# STYLE HEADER
# =====================================================

def style_header(sheet):

    for cell in sheet[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = thin_border


# =====================================================
# CREATE BRANCH EXCEL FILE
# =====================================================

def create_excel_file(class_name):

    file_path = get_excel_file(
        class_name
    )

    if not os.path.exists(file_path):

        workbook = Workbook()

        # =================================================
        # ATTENDANCE SHEET
        # =================================================

        attendance = workbook.active

        attendance.title = "Attendance"

        attendance.append([
            "Date",
            "Time",
            "Class",
            "Roll No",
            "Student Name",
            "Gender",
            "Status"
        ])

        style_header(
            attendance
        )


        # =================================================
        # DAILY REPORT
        # =================================================

        daily = workbook.create_sheet(
            "Daily Report"
        )

        daily.append([
            "Date",
            "Present",
            "Absent",
            "Total",
            "Attendance %"
        ])

        style_header(
            daily
        )


        # =================================================
        # WEEKLY REPORT
        # =================================================

        weekly = workbook.create_sheet(
            "Weekly Report"
        )

        weekly.append([
            "Week",
            "Start Date",
            "End Date",
            "Present",
            "Absent",
            "Total",
            "Attendance %"
        ])

        style_header(
            weekly
        )


        # =================================================
        # MONTHLY REPORT
        # =================================================

        monthly = workbook.create_sheet(
            "Monthly Report"
        )

        monthly.append([
            "Month",
            "Present",
            "Absent",
            "Total",
            "Attendance %"
        ])

        style_header(
            monthly
        )


        workbook.save(
            file_path
        )

        print(
            f"{class_name} Excel file created!"
        )


    return file_path


# =====================================================
# UPDATE REPORT SHEETS
# =====================================================

def update_reports(class_name):

    file_path = create_excel_file(
        class_name
    )

    workbook = load_workbook(
        file_path
    )

    attendance_sheet = workbook[
        "Attendance"
    ]

    daily_sheet = workbook[
        "Daily Report"
    ]

    weekly_sheet = workbook[
        "Weekly Report"
    ]

    monthly_sheet = workbook[
        "Monthly Report"
    ]


    # =================================================
    # READ ATTENDANCE
    # =================================================

    records = []

    for row in attendance_sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        if not row[0]:

            continue

        records.append({

            "date": row[0],

            "time": row[1],

            "className": row[2],

            "rollNo": row[3],

            "studentName": row[4],

            "gender": row[5],

            "status": row[6]

        })


    # =================================================
    # DAILY REPORT
    # =================================================

    daily_data = {}

    for record in records:

        date = record["date"]

        if date not in daily_data:

            daily_data[date] = {

                "present": 0,

                "absent": 0

            }


        if record["status"] == "Present":

            daily_data[
                date
            ]["present"] += 1


        elif record["status"] == "Absent":

            daily_data[
                date
            ]["absent"] += 1


    if daily_sheet.max_row > 1:

        daily_sheet.delete_rows(
            2,
            daily_sheet.max_row
        )


    for date, data in sorted(
        daily_data.items()
    ):

        present = data["present"]

        absent = data["absent"]

        total = present + absent

        percentage = (

            round(
                (present / total) * 100,
                2
            )

            if total > 0

            else 0

        )

        daily_sheet.append([

            date,

            present,

            absent,

            total,

            percentage

        ])


    # =================================================
    # WEEKLY REPORT
    # =================================================

    weekly_data = {}

    for record in records:

        try:

            date_obj = datetime.strptime(
                record["date"],
                "%d-%m-%Y"
            )

        except:

            continue


        monday = (

            date_obj

            - timedelta(
                days=date_obj.weekday()
            )

        )

        sunday = (

            monday

            + timedelta(days=6)

        )

        week_key = monday.strftime(
            "%d-%m-%Y"
        )


        if week_key not in weekly_data:

            weekly_data[week_key] = {

                "start": monday,

                "end": sunday,

                "present": 0,

                "absent": 0

            }


        if record["status"] == "Present":

            weekly_data[
                week_key
            ]["present"] += 1


        elif record["status"] == "Absent":

            weekly_data[
                week_key
            ]["absent"] += 1


    if weekly_sheet.max_row > 1:

        weekly_sheet.delete_rows(
            2,
            weekly_sheet.max_row
        )


    for week, data in sorted(
        weekly_data.items()
    ):

        present = data["present"]

        absent = data["absent"]

        total = present + absent

        percentage = (

            round(
                (present / total) * 100,
                2
            )

            if total > 0

            else 0

        )


        weekly_sheet.append([

            week,

            data["start"].strftime(
                "%d-%m-%Y"
            ),

            data["end"].strftime(
                "%d-%m-%Y"
            ),

            present,

            absent,

            total,

            percentage

        ])


    # =================================================
    # MONTHLY REPORT
    # =================================================

    monthly_data = {}

    for record in records:

        try:

            date_obj = datetime.strptime(
                record["date"],
                "%d-%m-%Y"
            )

        except:

            continue


        month_key = date_obj.strftime(
            "%m-%Y"
        )


        if month_key not in monthly_data:

            monthly_data[month_key] = {

                "present": 0,

                "absent": 0

            }


        if record["status"] == "Present":

            monthly_data[
                month_key
            ]["present"] += 1


        elif record["status"] == "Absent":

            monthly_data[
                month_key
            ]["absent"] += 1


    if monthly_sheet.max_row > 1:

        monthly_sheet.delete_rows(
            2,
            monthly_sheet.max_row
        )


    for month, data in sorted(
        monthly_data.items()
    ):

        present = data["present"]

        absent = data["absent"]

        total = present + absent

        percentage = (

            round(
                (present / total) * 100,
                2
            )

            if total > 0

            else 0

        )


        month_obj = datetime.strptime(
            month,
            "%m-%Y"
        )

        month_name = month_obj.strftime(
            "%B %Y"
        )


        monthly_sheet.append([

            month_name,

            present,

            absent,

            total,

            percentage

        ])


    # =================================================
    # AUTO COLUMN WIDTH
    # =================================================

    for sheet in [

        attendance_sheet,

        daily_sheet,

        weekly_sheet,

        monthly_sheet

    ]:

        for column in sheet.columns:

            max_length = 0

            column_letter = (
                column[0].column_letter
            )

            for cell in column:

                if cell.value:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )


            sheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                35
            )


        # =================================================
        # BORDERS
        # =================================================

        for row in sheet.iter_rows():

            for cell in row:

                cell.border = thin_border


    # =================================================
    # SAVE
    # =================================================

    workbook.save(
        file_path
    )


# =====================================================
# HOME
# =====================================================

@app.route(
    "/",
    methods=["GET"]
)
def home():

    return jsonify({

        "success": True,

        "message":
        "College Attendance Backend is running!"

    })


# =====================================================
# TEACHER LOGIN
# =====================================================

@app.route("/api/teacher-login", methods=["POST"])
def teacher_login():

    try:

        data = request.get_json()

        if not data:
            return jsonify({
                "success": False,
                "message": "Login data is required"
            }), 400

        email = str(
            data.get("email", "")
        ).strip().lower()

        password = str(
            data.get("password", "")
        ).strip()

        print("LOGIN EMAIL:", email)

        # ==============================================
        # CHECK TEACHER LOGIN
        # ==============================================

        for teacher in TEACHERS:

            if (
                email == teacher["email"].lower()
                and password == teacher["password"]
            ):

                print(
                    "LOGIN SUCCESS:",
                    email
                )

                return jsonify({

                    "success": True,

                    "message": "Login successful",

                    "teacher": {

                        "email": teacher["email"],

                        "name": teacher["name"],

                        "role": teacher["role"]

                    }

                }), 200

        # ==============================================
        # INVALID LOGIN
        # ==============================================

        print(
            "INVALID LOGIN:",
            email
        )

        return jsonify({

            "success": False,

            "message": "Invalid email or password"

        }), 401

    except Exception as error:

        print(
            "Login Error:",
            error
        )

        return jsonify({

            "success": False,

            "message": "Login server error"

        }), 500


# =====================================================
# MARK ATTENDANCE
# =====================================================

@app.route(
    "/api/attendance",
    methods=["POST"]
)
def mark_attendance():

    try:

        data = request.get_json()


        class_name = data.get(
            "className"
        )

        roll_no = data.get(
            "rollNo"
        )

        student_name = data.get(
            "studentName"
        )

        gender = data.get(
            "gender"
        )

        status = data.get(
            "status"
        )


        # =================================================
        # VALIDATION
        # =================================================

        if not class_name:

            return jsonify({

                "success": False,

                "message":
                "Class name is required"

            }), 400


        if not roll_no:

            return jsonify({

                "success": False,

                "message":
                "Roll number is required"

            }), 400


        if not student_name:

            return jsonify({

                "success": False,

                "message":
                "Student name is required"

            }), 400


        if status not in [

            "Present",

            "Absent"

        ]:

            return jsonify({

                "success": False,

                "message":
                "Invalid attendance status"

            }), 400


        # =================================================
        # CREATE BRANCH FILE
        # =================================================

        file_path = create_excel_file(
            class_name
        )


        workbook = load_workbook(
            file_path
        )

        sheet = workbook[
            "Attendance"
        ]


        # =================================================
        # DATE & TIME
        # =================================================

        now = datetime.now()

        current_date = now.strftime(
            "%d-%m-%Y"
        )

        current_time = now.strftime(
            "%H:%M:%S"
        )


        # =================================================
        # CHECK DUPLICATE
        # =================================================

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            if (

                row[0] == current_date

                and str(row[3])
                == str(roll_no)

                and row[2]
                == class_name

            ):

                workbook.close()

                return jsonify({

                    "success": False,

                    "message":
                    "Attendance already marked for today"

                }), 409


        # =================================================
        # SAVE ATTENDANCE
        # =================================================

        sheet.append([

            current_date,

            current_time,

            class_name,

            roll_no,

            student_name,

            gender,

            status

        ])


        workbook.save(
            file_path
        )

        workbook.close()


        # =================================================
        # UPDATE REPORTS
        # =================================================

        update_reports(
            class_name
        )


        print(

            f"Attendance saved: "

            f"{student_name} - "

            f"{class_name} - "

            f"{status}"

        )


        return jsonify({

            "success": True,

            "message":
            "Attendance saved successfully!",

            "student":
            student_name,

            "class":
            class_name,

            "rollNo":
            roll_no,

            "status":
            status

        }), 200


    except Exception as error:

        print(
            "Attendance Error:",
            error
        )

        return jsonify({

            "success": False,

            "message":
            str(error)

        }), 500


# =====================================================
# GET ALL ATTENDANCE
# =====================================================

@app.route(
    "/api/attendance",
    methods=["GET"]
)
def get_attendance():

    try:

        attendance_data = []


        for file_name in os.listdir(
            EXCEL_FOLDER
        ):

            if not file_name.endswith(
                "_Attendance.xlsx"
            ):

                continue


            file_path = os.path.join(
                EXCEL_FOLDER,
                file_name
            )


            workbook = load_workbook(
                file_path,
                read_only=True
            )


            if (

                "Attendance"
                not in workbook.sheetnames

            ):

                workbook.close()

                continue


            sheet = workbook[
                "Attendance"
            ]


            for row in sheet.iter_rows(
                min_row=2,
                values_only=True
            ):

                if not row[0]:

                    continue


                attendance_data.append({

                    "date": row[0],

                    "time": row[1],

                    "className": row[2],

                    "rollNo": row[3],

                    "studentName": row[4],

                    "gender": row[5],

                    "status": row[6]

                })


            workbook.close()


        return jsonify({

            "success": True,

            "attendance":
            attendance_data

        }), 200


    except Exception as error:

        print(
            "Attendance GET Error:",
            error
        )

        return jsonify({

            "success": False,

            "message":
            str(error)

        }), 500


# =====================================================
# DASHBOARD DATA
# =====================================================

@app.route(
    "/api/dashboard",
    methods=["GET"]
)
def dashboard_data():

    try:

        today = datetime.now().strftime(
            "%d-%m-%Y"
        )


        # =================================================
        # TOTAL STUDENTS
        # =================================================

        total_students = 180


        present = 0

        absent = 0


        # =================================================
        # READ ALL BRANCH FILES
        # =================================================

        for file_name in os.listdir(
            EXCEL_FOLDER
        ):

            if not file_name.endswith(
                "_Attendance.xlsx"
            ):

                continue


            file_path = os.path.join(
                EXCEL_FOLDER,
                file_name
            )


            workbook = load_workbook(
                file_path,
                read_only=True
            )


            if (

                "Attendance"
                not in workbook.sheetnames

            ):

                workbook.close()

                continue


            sheet = workbook[
                "Attendance"
            ]


            for row in sheet.iter_rows(
                min_row=2,
                values_only=True
            ):

                if row[0] != today:

                    continue


                if row[6] == "Present":

                    present += 1


                elif row[6] == "Absent":

                    absent += 1


            workbook.close()


        # =================================================
        # CALCULATE PERCENTAGE
        # =================================================

        total_marked = (
            present + absent
        )


        attendance_percentage = 0


        if total_marked > 0:

            attendance_percentage = round(

                (

                    present
                    / total_marked

                ) * 100,

                2

            )


        return jsonify({

            "success": True,

            "totalStudents":
            total_students,

            "present":
            present,

            "absent":
            absent,

            "attendancePercentage":
            attendance_percentage

        }), 200


    except Exception as error:

        print(
            "Dashboard Error:",
            error
        )

        return jsonify({

            "success": False,

            "message":
            str(error)

        }), 500


# =====================================================
# RUN SERVER
# =====================================================

if __name__ == "__main__":

    print(
        "---------------------------------------"
    )

    print(
        "College Attendance System"
    )

    print(
        "Backend Server Starting..."
    )

    print(
        "---------------------------------------"
    )


    app.run(

        host="127.0.0.1",

        port=5000,

        debug=True

    )